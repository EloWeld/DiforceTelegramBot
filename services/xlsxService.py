from etc.helpers import rdotdict
from loader import MDB
from services.oneService import OneService
import openpyxl
from openpyxl.styles import Font, Alignment
from services.oneService import OneService


class XLSXService:
    @staticmethod
    def generate_price_xlsx():
        catalog = [rdotdict(x) for x in list(MDB.Goods.find())]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Price List"

        # Заголовки столбцов
        headers = ["ID", "Name", "PriceЦБ0000001", "Quantity in Stores"]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Заполнение данными
        for row_num, item in enumerate(catalog, 2):
            sheet.cell(row=row_num, column=1, value=item["ProductID"])
            sheet.cell(row=row_num, column=2, value=item["ProductName"])
            sheet.cell(row=row_num, column=3, value=item["PriceЦБ0000001"])
            quantity_in_stores = ", ".join(
                [f"{x['store_name']}: {x['quantity']}" for x in item["QuantityInStores"]]
            )
            sheet.cell(row=row_num, column=4, value=quantity_in_stores)

        # Автоматическое изменение ширины столбцов
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column].width = max_length + 2

        # Сохранение файла
        file_name = "price_list.xlsx"
        workbook.save(file_name)

        return file_name